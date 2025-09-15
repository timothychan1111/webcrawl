import streamlit as st
import time
import random
import os
import pandas as pd
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook

# --------------------
# CONFIG
# --------------------
URLS = {
    "S&P": "https://finance.yahoo.com/quote/%5EGSPC/history?p=%5EGSPC",
    "DJI": "https://finance.yahoo.com/quote/%5EDJI/history?p=%5EDJI",
    "NAS": "https://finance.yahoo.com/quote/%5EIXIC/history?p=%5EIXIC",
    "RUT":"https://finance.yahoo.com/quote/%5ERUT/history?p=%5ERUT"
}

EXCEL_FILE = "s.xlsx"
END_DATE = datetime.today()
START_DATE = END_DATE - timedelta(days=30)

# --------------------
# STREAMLIT APP
# --------------------
st.title("Yahoo Finance Auto-Updater")
st.write("Fetches S&P500, Dow Jones, Nasdaq and updates Excel with newest data on top.")

if st.button("Fetch & Update Data"):

    st.info("Starting scraping process...")

    # --------------------
    # INIT SELENIUM DRIVER
    # --------------------
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)

    # --------------------
    # READ EXISTING EXCEL
    # --------------------
    if os.path.exists(EXCEL_FILE) and os.path.getsize(EXCEL_FILE) > 0:
        existing_sheets = pd.read_excel(EXCEL_FILE, sheet_name=None, engine="openpyxl")
    else:
        existing_sheets = {}

    # --------------------
    # HELPER FUNCTION
    # --------------------
    def fetch_table_data(url, start_date, end_date):
        driver.get(url)
        time.sleep(random.uniform(2, 4))
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(2, 4))

        table = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "table"))
        )
        rows = table.find_elements(By.TAG_NAME, "tr")
        data = []
        for row in rows[1:]:
            cols = row.find_elements(By.TAG_NAME, "td")
            if len(cols) < 6:
                continue
            try:
                date = datetime.strptime(cols[0].text, "%b %d, %Y")
            except:
                continue
            if not (start_date <= date <= end_date):
                continue
            open_price = cols[1].text.replace(",", "")
            high_price = cols[2].text.replace(",", "")
            low_price = cols[3].text.replace(",", "")
            close_price = cols[4].text.replace(",", "")
            adj_close = cols[5].text.replace(",", "")
            volume = cols[6].text.replace(",", "") if len(cols) > 6 else ""
            data.append([
                date.strftime("%Y/%m/%d"),
                open_price, high_price, low_price, close_price, adj_close, volume
            ])
            time.sleep(random.uniform(0.1, 0.3))
        return data

    # --------------------
    # FETCH & UPDATE
    # --------------------
    failed_urls = {}

    for name, url in URLS.items():
        st.write(f" Fetching {name} ...")
        try:
            data = fetch_table_data(url, START_DATE, END_DATE)
            if not data:
                st.warning(f" No data for {name}")
                failed_urls[name] = url
                continue

            df_new = pd.DataFrame(data, columns=["Date","Open","High","Low","Close","Adj Close","Volume"])
            if name in existing_sheets:
                df_existing = existing_sheets[name]
                df_combined = pd.concat([df_existing, df_new]).drop_duplicates(subset=["Date"]).sort_values("Date", ascending=False)
            else:
                df_combined = df_new.sort_values("Date", ascending=False)

            existing_sheets[name] = df_combined
            st.success(f" {name} updated ({len(df_combined)} rows)")
            time.sleep(random.uniform(2, 5))
        except Exception as e:
            st.error(f" Error fetching {name}: {e}")
            failed_urls[name] = url

    # --------------------
    # RETRY FAILED
    # --------------------
    if failed_urls:
        st.write("♻ Retrying failed URLs...")
        for name, url in failed_urls.items():
            try:
                data = fetch_table_data(url, START_DATE, END_DATE)
                if not data:
                    st.warning(f" Still no data for {name}, skipping.")
                    continue
                df_new = pd.DataFrame(data, columns=["Date","Open","High","Low","Close","Adj Close","Volume"])
                if name in existing_sheets:
                    df_existing = existing_sheets[name]
                    df_combined = pd.concat([df_existing, df_new]).drop_duplicates(subset=["Date"]).sort_values("Date", ascending=False)
                else:
                    df_combined = df_new.sort_values("Date", ascending=False)
                existing_sheets[name] = df_combined
                st.success(f"✅ {name} updated on retry ({len(df_combined)} rows)")
            except Exception as e:
                st.error(f" Failed again for {name}, skipping: {e}")

    # --------------------
    # SAVE TO EXCEL (BN–BT, row2 headers, row3+ data, newest on top)
    # --------------------
    def save_to_excel(existing_sheets, file_path):
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()

        for sheet_name, df in existing_sheets.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            start_col = 66  # BN
            start_row = 2   # headers row

            # Ensure column types
            df["Date"] = df["Date"].astype(str)
            for col in ["Open", "High", "Low", "Close", "Adj Close"]:
                df[col] = pd.to_numeric(df[col], errors="coerce").astype(float)
            df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce").astype("Int64")

            # Sort descending
            df = df.sort_values("Date", ascending=False).reset_index(drop=True)

            # Collect existing dates
            existing_dates = set()
            for r in ws.iter_rows(min_row=start_row+1, min_col=start_col, max_col=start_col, values_only=True):
                if r[0] is not None:
                    existing_dates.add(str(r[0]))

            # Write headers
            for i, col_name in enumerate(df.columns, start=start_col):
                ws.cell(row=start_row, column=i, value=col_name)

            # Insert missing rows at top (row 3)
            row_pointer = start_row + 1
            for _, row in df.iterrows():
                if row["Date"] in existing_dates:
                    continue
                ws.insert_rows(row_pointer)
                ws.cell(row=row_pointer, column=start_col, value=row["Date"])
                ws.cell(row=row_pointer, column=start_col+1, value=float(row["Open"]))
                ws.cell(row=row_pointer, column=start_col+2, value=float(row["High"]))
                ws.cell(row=row_pointer, column=start_col+3, value=float(row["Low"]))
                ws.cell(row=row_pointer, column=start_col+4, value=float(row["Close"]))
                ws.cell(row=row_pointer, column=start_col+5, value=float(row["Adj Close"]))
                vol_val = None if pd.isna(row["Volume"]) else int(row["Volume"])
                ws.cell(row=row_pointer, column=start_col+6, value=vol_val)

        wb.save(file_path)

    save_to_excel(existing_sheets, EXCEL_FILE)
    driver.quit()
    st.success(f"All data saved to {EXCEL_FILE}")
