import time
import random
import os
import pandas as pd
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook

# --------------------
# CONFIG
# --------------------
URLS = {
    "S&P500": "https://finance.yahoo.com/quote/%5EGSPC/history?p=%5EGSPC",
    "DowJones": "https://finance.yahoo.com/quote/%5EDJI/history?p=%5EDJI",
    "Nasdaq": "https://finance.yahoo.com/quote/%5EIXIC/history?p=%5EIXIC",
}

EXCEL_FILE = "yfdata_latest_month.xlsx"
END_DATE = datetime.today()
START_DATE = END_DATE - timedelta(days=30)

# --------------------
# INITIALIZE DRIVER (NO chromedriver PATH needed!)
# --------------------
options = webdriver.ChromeOptions()
# options.add_argument("--headless")   # uncomment if you don’t want browser window
driver = webdriver.Chrome(options=options)

# --------------------
# HELPER FUNCTION
# --------------------
def fetch_table_data(url, start_date, end_date):
    driver.get(url)
    time.sleep(random.uniform(2, 5))
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(random.uniform(2, 4))

    table = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.TAG_NAME, "table"))
    )
    rows = table.find_elements(By.TAG_NAME, "tr")
    data = []
    for row in rows[1:]:
        cols = row.find_elements(By.TAG_NAME, "td")
        if len(cols) < 6:
            continue
        try:
            date = datetime.strptime(cols[0].text, "%b %d, %Y")
        except:
            continue
        if not (start_date <= date <= end_date):
            continue
        open_price = cols[1].text.replace(",", "")
        high_price = cols[2].text.replace(",", "")
        low_price = cols[3].text.replace(",", "")
        close_price = cols[4].text.replace(",", "")
        adj_close = cols[5].text.replace(",", "")
        volume = cols[6].text.replace(",", "") if len(cols) > 6 else ""
        data.append([date, open_price, high_price, low_price, close_price, adj_close, volume])
        time.sleep(random.uniform(0.1, 0.3))
    return data

# --------------------
# FETCH DATA
# --------------------
all_data = {}
for name, url in URLS.items():
    print(f"\nFetching {name} ...")
    try:
        data = fetch_table_data(url, START_DATE, END_DATE)
        df_new = pd.DataFrame(data, columns=["Date","Open","High","Low","Close","Adj Close","Volume"])
        all_data[name] = df_new
        print(f"✅ {name} fetched {len(df_new)} rows")
    except Exception as e:
        print(f"❌ Error fetching {name}: {e}")

driver.quit()

# --------------------
# SAVE TO EXCEL (BN2 headers, BN3+ data)
# --------------------
if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
else:
    wb = Workbook()

for name, df in all_data.items():
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)

    start_col = 66  # BN column
    headers = list(df.columns)

    # Write headers in row 2
    for j, header in enumerate(headers):
        ws.cell(row=2, column=start_col + j, value=header)

    # Collect existing dates
    existing_dates = set()
    for row in ws.iter_rows(min_row=3, min_col=start_col, max_col=start_col, values_only=True):
        if row[0] is not None:
            try:
                existing_dates.add(pd.to_datetime(row[0]).date())
            except:
                continue

    # Append only missing rows
    for _, row in df.iterrows():
        d = row["Date"].date()
        if d in existing_dates:
            continue
        excel_row = ws.max_row + 1 if ws.max_row >= 3 else 3
        for j, val in enumerate(row):
            ws.cell(row=excel_row, column=start_col + j, value=val)

wb.save(EXCEL_FILE)
print(f"\nAll data written to {EXCEL_FILE} (each index in its own sheet, BN2–BT..)")
